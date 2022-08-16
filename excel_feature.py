import openpyxl
import pickle
from openpyxl.styles import Alignment

# map the column letter to corresponding day in 'blank_week' spreadsheet.
# days are hardcoded as they will never change.
day_columns = {'Sunday': 'C', 'Monday': 'D', 'Tuesday': 'E',
               'Wednesday': 'F', 'Thursday': 'G', 'Friday': 'H',
               'Saturday': 'I'}


def row_updater(row, row_dict, cut_off=''):
    """When colleague is added/removed/has department changed, the row numbers
       of the colleagues need to be adjusted. The parameters are the row of the
       colleague in question, the dictionary containing the mapped row numbers
       and an optional cut_off parameter that is supplied when a colleague
       changes department. When their current row is deleted, all rows after are
       automatically adjusted and don't need to be increased."""
    for key, value in row_dict.items():
        # if cut-off row number is supplied, we increase the other rows up until
        # this row.
        if cut_off:
            if row <= value < cut_off:
                row_dict[key] += 1
        else:
            if value >= row:
                row_dict[key] += 1


def col_to_excel(excel_file, col, row_dict, cutoff=''):
    """Adds a col to the work_sheet and returns the row number in the
       worksheet."""
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    new_row = row_dict[col.department] + 1
    # the values for the relevant rows are increased by 1.
    row_updater(new_row, row_dict, cutoff)
    # a blank row is inserted at the previously determined value.
    sheet.insert_rows(new_row)
    # the relevant info is added to the worksheet. Column A contains
    # the dept headings and therefore, the col's are placed under
    # the corresponding one. B is for the colleague hours.
    # The correct alignment is insured using the imported Alignment class from
    # openpyxl.styles.
    name_cell = 'A' + str(new_row)
    hours_cell = 'B' + str(new_row)
    sheet[name_cell] = col.name()
    sheet[name_cell].alignment = Alignment(horizontal="left")
    sheet[hours_cell] = col.hours
    sheet[hours_cell].alignment = Alignment(horizontal="left")
    wb.save(excel_file)
    wb.close()
    return new_row


def date_adder(worksheet, date_list):
    # Adds the dates of each day of the week to the worksheet, directly below
    # the day itself.
    wb = openpyxl.load_workbook(worksheet)
    sheet = wb.active
    # isolate the cells where the dates will be added.
    for row_of_cells in sheet['C3':'I3']:
        # iterate through the row itself, tracking the index to match up with
        # the relevant date, and the cell object.
        for index, cell in enumerate(row_of_cells):
            # isolate the coordinate value from the cell object for easier
            # assignment of value.
            coord = cell.coordinate
            # assign the date to the relevant cell in the format of day (as
            # zero-padded decimal number) and month (abbreviated name).
            sheet[coord] = date_list[index].strftime('%d-%b')
    output_file = f'Colleague Rota {date_list[0]}.xlsx'
    wb.save(output_file)
    wb.close()
    # return the filename for the updated spreadsheet.
    return output_file


def add_to_worksheet(worksheet, col_name, day, shift=''):
    """
    Adds final shift for colleague to the supplied spreadsheet.The coordinate
    corresponding to the colleague is determined by combining the column
    corresponding to the particular day with a string of the row number of the
    colleague
    """
    # open the desired spreadsheet and determine the active sheet.
    wb = openpyxl.load_workbook(worksheet)
    sheet = wb.active
    with open('ws_rows', 'rb') as f:
        ws_rows = pickle.load(f)
    # the corresponding row and column are assigned to variables for easier
    # combination later on.
    column = day_columns[day]
    row = str(ws_rows[col_name])
    if not shift:
        sheet[column + row] = 'O'
    else:
        sheet[column + row] = shift
    # the changes are saved and the spreadsheet is closed.
    wb.save(worksheet)
    wb.close()
    return None
